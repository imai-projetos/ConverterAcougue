import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import tempfile
import os

# ⛳ Deve ser o primeiro comando do Streamlit
st.set_page_config(page_title="Formatador de Inventário", page_icon="📊", layout="centered")

# Logo da empresa
st.image(r"https://github.com/imai-projetos/ConverterAcougue/blob/main/imaiempresas_logo.jpeg?raw=true", width=200)

# Título do app
st.title("📄 Formatador de Inventário do Açougue em Excel")

def remover_mesclas_e_formatar(caminho_entrada, caminho_saida):
    wb = openpyxl.load_workbook(caminho_entrada)

    if "Document map" in wb.sheetnames:
        std = wb["Document map"]
        wb.remove(std)

    for ws in wb.worksheets:
        # 1. Remover mesclagens
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))

        # 2. Excluir colunas vazias
        max_col = ws.max_column
        colunas_para_excluir = []

        for col in range(1, max_col + 1):
            vazia = True
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
                if row[0] not in (None, '', ' '):
                    vazia = False
                    break
            if vazia:
                colunas_para_excluir.append(col)

        for col_idx in sorted(colunas_para_excluir, reverse=True):
            ws.delete_cols(col_idx)

        # 3. Excluir linhas de 1 a 14
        ws.delete_rows(1, 14)

        # 4. Excluir colunas fixas
        colunas_excluir = ['AB', 'AA', 'X', 'U', 'T', 'R', 'M', 'L', 'J', 'H', 'F', 'C']
        for letra in sorted(colunas_excluir, key=column_index_from_string, reverse=True):
            idx = column_index_from_string(letra)
            if idx <= ws.max_column:
                ws.delete_cols(idx)

        # 5. Remover imagens
        if hasattr(ws, "_images"):
            ws._images.clear()

        # 6. Remover linhas onde a coluna B está vazia
        col_B_idx = column_index_from_string('B')
        for row in reversed(list(ws.iter_rows(min_row=1))):
            if len(row) >= col_B_idx:
                valor = row[col_B_idx - 1].value
                if valor in (None, '', ' '):
                    ws.delete_rows(row[0].row)

        # 7. Mostrar linhas de grade
        ws.sheet_view.showGridLines = True

        # 8. Descongelar painéis
        ws.freeze_panes = None

        # 9. Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # 10. Ajustar altura das linhas
        for row in ws.iter_rows():
            max_height = 15
            for cell in row:
                if cell.value:
                    lines = str(cell.value).count('\n') + 1
                    max_height = max(max_height, lines * 15)
            ws.row_dimensions[row[0].row].height = max_height

    wb.save(caminho_saida)


# UPLOAD E PROCESSAMENTO
arquivo = st.file_uploader("Envie o arquivo Excel (.xlsx):", type=["xlsx"])

if arquivo is not None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_input:
        tmp_input.write(arquivo.read())
        tmp_input.flush()

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_output:
            remover_mesclas_e_formatar(tmp_input.name, tmp_output.name)

            with open(tmp_output.name, "rb") as f:
                st.success("✅ Arquivo processado com sucesso!")
                st.download_button(
                    label="📥 Baixar Arquivo Formatado",
                    data=f.read(),
                    file_name=f"formatado_{arquivo.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    os.remove(tmp_input.name)